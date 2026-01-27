#!/usr/bin/env python3
"""
IA Data Extractor Web Application
Simple web interface for extracting IA form data from PDFs.
"""

import os
import re
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename

# Try importing PDF libraries - catch errors for debugging
try:
    import pdfplumber
    PDF_IMPORT_ERROR = None
except Exception as e:
    pdfplumber = None
    PDF_IMPORT_ERROR = str(e)

try:
    from openpyxl import Workbook
    EXCEL_IMPORT_ERROR = None
except Exception as e:
    Workbook = None
    EXCEL_IMPORT_ERROR = str(e)

app = Flask(__name__)

# Use /tmp for cloud deployments (Render uses ephemeral filesystem)
if os.environ.get('RENDER'):
    app.config['UPLOAD_FOLDER'] = Path('/tmp/uploads')
    app.config['OUTPUT_FOLDER'] = Path('/tmp/output')
else:
    app.config['UPLOAD_FOLDER'] = Path(__file__).parent / 'uploads'
    app.config['OUTPUT_FOLDER'] = Path(__file__).parent / 'output'

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'pdf'}

# Ensure folders exist
app.config['UPLOAD_FOLDER'].mkdir(parents=True, exist_ok=True)
app.config['OUTPUT_FOLDER'].mkdir(parents=True, exist_ok=True)


def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


class IADataExtractor:
    """Extracts data from IA PDF forms."""

    def extract_text_from_page(self, page):
        """Extract text from a single PDF page."""
        return page.extract_text()

    def parse_burial_number(self, text):
        """Extract burial number from form."""
        match = re.search(r'Burial Number\s+Contract Number\s*\n\s*(\d+)', text)
        if match:
            return match.group(1).strip()
        return ""

    def parse_deceased_name(self, text):
        """Extract deceased name in 'Last, First' format."""
        match = re.search(r'Last Name First Name MI Suffix.*?\n(.+?)\n', text)
        if match:
            data_line = match.group(1).strip()
            parts = data_line.split()
            if len(parts) >= 2:
                last_name = parts[0]
                first_name = parts[1]
                return f"{last_name}, {first_name}"
        return ""

    def parse_date_of_death(self, text):
        """Extract and format date of death as MM/DD/YYYY."""
        match = re.search(r'Last Name First Name MI Suffix.*?\n(.+?)\n', text)
        if match:
            data_line = match.group(1).strip()
            date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})\s+(?:Male|Female)', data_line)
            if date_match:
                date_str = date_match.group(1)
                parts = date_str.split('/')
                if len(parts) == 3:
                    mm, dd, yyyy = parts
                    return f"{mm.zfill(2)}/{dd.zfill(2)}/{yyyy}"
        return ""

    def parse_nok_name(self, text):
        """Extract NOK (Next of Kin) name in 'Last, First' format."""
        match = re.search(r'DATA ON NEXT OF KIN OR REPRESENTATIVE.*?Last Name First Name MI Suffix.*?\n(.+?)\n', text, re.DOTALL)
        if match:
            nok_line = match.group(1).strip()
            parts = nok_line.split()
            if len(parts) >= 2:
                last_name = parts[0]
                first_name = parts[1]
                return f"{last_name}, {first_name}"
        return ""

    def parse_relationship(self, text):
        """Extract relationship to deceased."""
        match = re.search(r'Relationship E-Mail Address:\s*\n(\S+)', text)
        if match:
            return match.group(1).strip()
        return ""

    def parse_phone(self, text):
        """Extract and format phone number as (###) ###-####."""
        match = re.search(r'DATA ON NEXT OF KIN OR REPRESENTATIVE.*?(\(\d{3}\)\s*\d{3}-\d{4})', text, re.DOTALL)
        if match:
            return match.group(1).strip()
        return ""

    def extract_ia_data(self, page):
        """Extract all IA data from a single page."""
        text = self.extract_text_from_page(page)

        data = {
            'burial_number': self.parse_burial_number(text),
            'deceased_name': self.parse_deceased_name(text),
            'dod': self.parse_date_of_death(text),
            'nok_name': self.parse_nok_name(text),
            'relationship': self.parse_relationship(text),
            'phone': self.parse_phone(text)
        }

        return data

    def process_pdf(self, pdf_path):
        """Process PDF and extract all IA data."""
        all_data = []

        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                data = self.extract_ia_data(page)
                all_data.append(data)

        return all_data

    def create_excel(self, data, output_path):
        """Create Excel file with extracted data."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "IA Log"

        # Write headers
        headers = ['Burial #', 'Deceased Name\n(Last, First)', 'DOD',
                   'NOK Name\n(Last, First)', 'Relation to\nDeceased', 'Contact Info']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Write data starting at row 2
        for idx, record in enumerate(data, start=2):
            sheet.cell(row=idx, column=1, value=record['burial_number'])
            sheet.cell(row=idx, column=2, value=record['deceased_name'])
            sheet.cell(row=idx, column=3, value=record['dod'])
            sheet.cell(row=idx, column=4, value=record['nok_name'])
            sheet.cell(row=idx, column=5, value=record['relationship'])
            sheet.cell(row=idx, column=6, value=record['phone'])

        workbook.save(output_path)


@app.route('/')
def index():
    """Render the main page."""
    return render_template('index.html')


@app.route('/health')
def health():
    """Health check endpoint for debugging."""
    status = {
        'app': 'running',
        'pdfplumber': 'ok' if pdfplumber else f'FAILED: {PDF_IMPORT_ERROR}',
        'openpyxl': 'ok' if Workbook else f'FAILED: {EXCEL_IMPORT_ERROR}',
        'upload_folder': str(app.config['UPLOAD_FOLDER']),
        'output_folder': str(app.config['OUTPUT_FOLDER']),
    }
    return jsonify(status)


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and processing."""
    try:
        # Check if libraries loaded correctly
        if PDF_IMPORT_ERROR:
            return jsonify({'error': f'PDF library failed to load: {PDF_IMPORT_ERROR}'}), 500
        if EXCEL_IMPORT_ERROR:
            return jsonify({'error': f'Excel library failed to load: {EXCEL_IMPORT_ERROR}'}), 500

        # Ensure folders exist
        app.config['UPLOAD_FOLDER'].mkdir(parents=True, exist_ok=True)
        app.config['OUTPUT_FOLDER'].mkdir(parents=True, exist_ok=True)

        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': 'Only PDF files are allowed'}), 400

        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        upload_path = app.config['UPLOAD_FOLDER'] / f"{timestamp}_{filename}"

        try:
            file.save(upload_path)
        except Exception as e:
            return jsonify({'error': f'Failed to save file: {str(e)}'}), 500

        # Process PDF
        try:
            extractor = IADataExtractor()
            data = extractor.process_pdf(upload_path)
        except Exception as e:
            upload_path.unlink(missing_ok=True)
            return jsonify({'error': f'Failed to process PDF: {str(e)}'}), 500

        # Create Excel file
        output_filename = f"IA_Log_{timestamp}.xlsx"
        output_path = app.config['OUTPUT_FOLDER'] / output_filename

        try:
            extractor.create_excel(data, output_path)
        except Exception as e:
            upload_path.unlink(missing_ok=True)
            return jsonify({'error': f'Failed to create Excel: {str(e)}'}), 500

        # Clean up uploaded PDF
        upload_path.unlink(missing_ok=True)

        return jsonify({
            'success': True,
            'filename': output_filename,
            'record_count': len(data)
        })

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"UPLOAD ERROR: {error_details}")
        return jsonify({'error': f'{type(e).__name__}: {str(e)}'}), 500


@app.route('/download/<filename>')
def download_file(filename):
    """Download the generated Excel file."""
    try:
        file_path = app.config['OUTPUT_FOLDER'] / filename
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@app.route('/cleanup')
def cleanup_old_files():
    """Clean up old files (optional maintenance endpoint)."""
    try:
        # Clean files older than 1 hour
        import time
        current_time = time.time()

        for folder in [app.config['UPLOAD_FOLDER'], app.config['OUTPUT_FOLDER']]:
            for file_path in folder.glob('*'):
                if file_path.is_file():
                    file_age = current_time - file_path.stat().st_mtime
                    if file_age > 3600:  # 1 hour
                        file_path.unlink()

        return jsonify({'success': True, 'message': 'Cleanup completed'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Get port from environment variable (for cloud deployment) or use 5001 (for local)
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=True, host='0.0.0.0', port=port)
